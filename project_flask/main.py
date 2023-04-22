# Step 2: Define the Flask application and routes
# - Import necessary modules
# - Initialize the Flask application
# - Create routes for the Home, Sign-in, and Dashboard pages

from flask import Flask, render_template, request, redirect, url_for, session
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.utils import secure_filename

app = Flask(__name__)


def style_excel_file(file):
    # Load the workbook and select the sheet you want to style
    wb = load_workbook(file)
    ws = wb.active

    # Increase font size for all cells
    font_size = 14
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=font_size)

    # Adjust column widths based on the length of the content
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 10.2
        ws.column_dimensions[column].width = adjusted_width

    # Adjust row heights based on the font size
    row_height = font_size * 2
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = row_height

    # Save the styled workbook
    wb.save(file)
    return


@app.route('/')
def home():
    return render_template('home.html')


UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
database_filename = "database.xlsx"


@app.route('/signin', methods=['GET', 'POST'])
def signin():
    error_message = None

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        database_filename = "database.xlsx"

        if os.path.exists(database_filename):
            database = pd.read_excel(database_filename)

            # Check if 'username' column exists before accessing it
            user = database.loc[database['username'] ==
                                username] if 'username' in database.columns else pd.DataFrame()

            # Check if 'password' column exists before accessing it
            if not user.empty and 'password' in user.columns and user['password'].iloc[0] == password:
                return render_template('dashboard.html', username=username)
            else:
                error_message = "Invalid username or password."
        else:
            error_message = "No users found in the database."

    return render_template('signin.html', error_message=error_message)
    style_excel_file(database_filename)


@app.route('/register', methods=['GET', 'POST'])
def register():
    error_message = None

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        database_filename = "database.xlsx"

        if os.path.exists(database_filename):
            database = pd.read_excel(database_filename)

            # Check if 'username' column exists before accessing it
            existing_user = database.loc[database['username'] ==
                                         username] if 'username' in database.columns else pd.DataFrame()

            if not existing_user.empty:
                error_message = "Username already exists."
            else:
                # Register the user by adding a new row to the DataFrame
                new_user = pd.DataFrame(
                    {"username": [username], "password": [password]})
                database = pd.concat([database, new_user], ignore_index=True)

                # Save the updated DataFrame to the Excel file
                database.to_excel(database_filename, index=False)

                return render_template('signin.html')
        else:
            # Create a new DataFrame with the new user's information and save it to the Excel file
            database = pd.DataFrame(
                {"username": [username], "password": [password]})
            database.to_excel(database_filename, index=False)

            return render_template('signin.html')

    return render_template('register.html', error_message=error_message)
    style_excel_file(database_filename)


# Route for the Dashboard page
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
database = "Enrollment.xlsx"


@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if request.method == 'POST':
        selected_option = request.form.get('selected_option')
        uploaded_file = request.files['uploaded_file']
        username = request.form.get('username')

        if uploaded_file and selected_option:
            filename = secure_filename(uploaded_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(file_path)

            file_info = pd.DataFrame({
                'Username': [username],
                'Course': [selected_option],
                'Resume': [filename],
            })

            df = pd.read_excel(database)
            new_row = pd.DataFrame(file_info, index=[0])
            df = pd.concat([df, new_row], ignore_index=True)
            df.to_excel(database, index=False)
            return render_template('dashboard.html')
        else:
            error_message = 'Error: File not uploaded or course not selected'
            return
    else:
        return render_template('dashboard.html')


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000, debug=True)
# Step 8: Update the signin route in main.py
# - Handle form submission
# - Save user details to an Excel file as a database
# - Redirect to the Dashboard page
# ...
